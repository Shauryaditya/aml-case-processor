# app/parser.py

from pathlib import Path
import csv
import pdfplumber
import re

DATE_RE = re.compile(r"\d{4}-\d{2}-\d{2}")
KNOWN_CHANNELS = {
    "ach", "wire", "cash", "atm", "card", "p2p", "crypto", "check"
}

def extract_transactions(file_path: str):
    path = Path(file_path)
    ext = path.suffix.lower()

    if ext == ".csv":
        return _extract_from_csv(path)
    elif ext in [".xlsx", ".xls"]:
        return _extract_from_excel(path)
    elif ext == ".pdf":
        return _extract_from_pdf(path)
    else:
        return []


def _extract_from_csv(path: Path):
    txs = []
    with path.open("r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            txs.append({
                "Date": row.get("Date") or row.get("date"),
                "amount": row.get("amount") or row.get("Amount"),
                "Type": row.get("Type") or row.get("type"),
                "Details": row.get("Details") or row.get("description") or "",
                "direction": row.get("Direction") or row.get("direction") or "unknown",
            })
    return txs


def _extract_from_excel(path: Path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    txs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {headers[i]: row[i] for i in range(len(headers))}
        txs.append({
            "Date": row_dict.get("Date") or row_dict.get("date"),
            "amount": row_dict.get("amount") or row_dict.get("Amount"),
            "Type": row_dict.get("Type") or row_dict.get("type"),
            "Details": row_dict.get("Details") or row_dict.get("description") or "",
            "direction": row_dict.get("Direction") or row_dict.get("direction") or "unknown",
        })
    return txs


def infer_direction_from_details(details: str) -> str:
    d = details.lower()

    inbound = (
        "incoming", "from ", "credit", "deposit", "salary", "payroll", "received"
    )
    outbound = (
        "transfer to", "wire to", "withdrawal", "payment", "sent", "debit", "purchase"
    )

    if any(k in d for k in inbound):
        return "inbound"
    if any(k in d for k in outbound):
        return "outbound"

    return "unknown"


def _extract_from_pdf(path: Path):
    """
    Handles three PDF formats:
    1. Complex_AML_Case: "Date Amount Type Direction Details" (1 amt with $, explicit direction)
    2. Mixed_200_Cases: "Date Description Channel Debit Credit" (no $, 2 amt columns)
    3. Upgraded_Business/Personal: "Date Description Channel Debit Credit Balance" (with $, 3 amt columns)
    """
    txs = []
    
    # Regex for amounts with dollar signs (including negative balances)
    money_with_dollar = re.compile(r"\$-?[\d,]+(?:\.\d{2})?")
    money_without_dollar = re.compile(r"(?<!\$)\b(\d{1,3}(?:,?\d{3})*(?:\.\d{2})?)\b")

    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            
            for line in text.splitlines():
                line = line.strip()

                # Skip headers and noise
                if not line or not DATE_RE.match(line[:10]):
                    continue
                if any(x in line.lower() for x in [
                    "opening balance", "closing balance", "statement period", 
                    "date amount type", "date description channel", "debit credit balance"
                ]):
                    continue

                date_str = line[:10]
                remaining = line[10:].strip()
                
                # Try extracting dollar amounts
                amounts = money_with_dollar.findall(remaining)
                has_dollar = len(amounts) > 0
                
                # If no dollar signs, try without
                if not has_dollar:
                    tokens = remaining.split()
                    amounts = [t for t in tokens if money_without_dollar.match(t)]
                
                if not amounts:
                    continue

                transaction_amount = None
                direction = "unknown"
                channel = "unknown"
                description = ""

                # Check if this is Complex_AML_Case format (has explicit direction keyword)
                has_explicit_direction = "inbound" in remaining.lower() or "outbound" in remaining.lower()

                if has_explicit_direction:
                    # FORMAT 1: Complex_AML_Case (Date Amount Type Direction Details)
                    transaction_amount = amounts[0]
                    
                    if "inbound" in remaining.lower():
                        direction = "inbound"
                    elif "outbound" in remaining.lower():
                        direction = "outbound"
                    
                    # Extract channel and description
                    tokens = remaining.split()
                    for i, t in enumerate(tokens):
                        if t.lower() in KNOWN_CHANNELS:
                            channel = t.lower()
                            # Find direction keyword position
                            dir_idx = -1
                            for j in range(i+1, len(tokens)):
                                if tokens[j].lower() in ["inbound", "outbound"]:
                                    dir_idx = j
                                    break
                            if dir_idx > 0 and dir_idx < len(tokens) - 1:
                                description = " ".join(tokens[dir_idx+1:]).lower()
                            break
                
                elif has_dollar and len(amounts) >= 2:
                    # CRITICAL: Identify which amounts are Debit/Credit vs Balance
                    # Strategy: Balance is typically the LAST amount and often much larger or negative
                    # The transaction amount is one of the first 1-2 amounts
                    
                    if len(amounts) == 2:
                        # FORMAT: Could be Debit+Credit OR Amount+Balance
                        # Heuristic: if second is much larger or negative, it's balance
                        first_val = float(amounts[0].replace('$', '').replace(',', ''))
                        second_val = float(amounts[1].replace('$', '').replace(',', ''))
                        
                        # If second is negative or 3x+ larger, it's likely the balance
                        if second_val < 0 or abs(second_val) > abs(first_val) * 3:
                            transaction_amount = amounts[0]
                            # Direction based on context since we don't have both debit/credit
                            direction = "outbound" if first_val > 0 else "inbound"
                        else:
                            # Both are transaction amounts (debit and credit columns)
                            # Pick non-zero one
                            if first_val > 0:
                                transaction_amount = amounts[0]
                                direction = "outbound"
                            elif second_val > 0:
                                transaction_amount = amounts[1]
                                direction = "inbound"
                    
                    elif len(amounts) == 3:
                        # FORMAT: Debit, Credit, Balance
                        # Last one is balance, first two are debit/credit
                        debit_str = amounts[0]
                        credit_str = amounts[1]
                        balance_str = amounts[2]  # Ignore this
                        
                        debit_val = float(debit_str.replace('$', '').replace(',', ''))
                        credit_val = float(credit_str.replace('$', '').replace(',', ''))
                        
                        # Credit (inbound) takes precedence
                        if credit_val > 0:
                            transaction_amount = credit_str
                            direction = "inbound"
                        elif debit_val > 0:
                            transaction_amount = debit_str
                            direction = "outbound"
                    
                    # Extract channel and description - BEFORE the amounts
                    # Remove all amounts from the line first
                    clean = remaining
                    for a in amounts:
                        clean = clean.replace(a, "", 1)  # Remove first occurrence
                    
                    tokens = clean.split()
                    
                    # Channel is typically the last meaningful token before amounts
                    for t in reversed(tokens):
                        if t.lower() in KNOWN_CHANNELS:
                            channel = t.lower()
                            break
                    
                    # Description is everything before channel
                    desc_tokens = []
                    for t in tokens:
                        if t.lower() == channel:
                            break
                        desc_tokens.append(t)
                    description = " ".join(desc_tokens).lower().strip()

                elif not has_dollar and len(amounts) >= 1:
                    # FORMAT 2: Mixed_200_Cases (no $)
                    transaction_amount = f"${amounts[0]}"
                    
                    # Extract channel
                    tokens = remaining.split()
                    for t in reversed(tokens):
                        if t.lower() in KNOWN_CHANNELS:
                            channel = t.lower()
                            break
                    
                    # Description
                    clean = remaining
                    for a in amounts:
                        clean = clean.replace(a, "", 1)
                    desc_tokens = []
                    for t in clean.split():
                        if t.lower() == channel:
                            break
                        desc_tokens.append(t)
                    description = " ".join(desc_tokens).lower().strip()
                    
                    # Infer direction
                    direction = infer_direction_from_details(description)
                    if direction == "unknown":
                        direction = "outbound"

                else:
                    # Fallback
                    transaction_amount = amounts[0] if amounts[0].startswith('$') else f"${amounts[0]}"
                    direction = infer_direction_from_details(remaining)
                    
                    tokens = remaining.split()
                    for t in reversed(tokens):
                        if t.lower() in KNOWN_CHANNELS:
                            channel = t.lower()
                            break
                    
                    description = " ".join(t for t in tokens if t.lower() != channel).lower()

                if not transaction_amount:
                    continue

                txs.append({
                    "Date": date_str,
                    "amount": transaction_amount,
                    "Type": channel,
                    "Details": description.strip(),
                    "direction": direction,
                })

    return txs